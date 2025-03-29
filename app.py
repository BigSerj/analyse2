# This file is deprecated. All functionality moved to appiframe.py

import os
from flask import Flask, render_template, request, send_file, jsonify, abort, Response
from markupsafe import Markup
import requests
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime, timedelta
import json
import threading
import math
from time import sleep, time
from openpyxl.styles.colors import Color
from copy import copy

app = Flask(__name__)

# Add port configuration
port = int(os.environ.get("PORT", 10000))

# Загрузка конфигурации
with open('config.py', 'r') as config_file:
    exec(config_file.read())

BASE_URL = 'https://api.moysklad.ru/api/remap/1.2'

processing_cancelled = False
processing_lock = threading.Lock()
current_status = {'total': 0, 'processed': 0}
api_request_times = []

# Добавляем корневой маршрут
@app.route('/')
def index():
    return render_template('iframe.html', 
                         stores=get_stores(),
                         product_groups=get_product_groups(),
                         product_groups_json=json.dumps(prepare_groups_for_js(get_product_groups())),
                         render_group_options=render_group_options)

# Добавляем маршрут iframe (теперь он будет дублировать корневой)
@app.route('/iframe')
def iframe():
    return index()

def prepare_groups_for_js(groups):
    result = []
    for group in groups:
        group_data = {
            'id': group['id'],
            'name': group['name'],
            'children': prepare_groups_for_js(group.get('children', [])),
            'hasChildren': bool(group.get('children'))
        }
        result.append(group_data)
    return result

@app.route('/process', methods=['POST'])
def process():
    try:
        global processing_cancelled, current_status, api_request_times
        with processing_lock:
            processing_cancelled = False
            current_status = {'total': 0, 'processed': 0}
            api_request_times = []
        
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        store_id = request.form['store_id']
        planning_days = int(request.form['planning_days'])
        
        product_groups = []
        if 'final_product_groups' in request.form:
            raw_groups = request.form.get('final_product_groups', '')
            if raw_groups:
                product_groups = [group.strip() for group in raw_groups.split(',') if group.strip()]
            print(f"Обработанные группы: {product_groups}")
        
        # Проверяем отмену перед получением данных
        if check_if_cancelled():
            return jsonify({'cancelled': True}), 200
            
        try:
            report_data = get_report_data(start_date, end_date, store_id, product_groups)
            print("Получены данные отчета:", report_data is not None)
            if report_data:
                print(f"Количество строк в отчете: {len(report_data.get('rows', []))}")
        except Exception as e:
            print(f"Ошибка при получении данных отчета: {str(e)}")
            import traceback
            print("Полный стек ошибки:")
            print(traceback.format_exc())
            return jsonify({'error': str(e)}), 500
        
        if not report_data or 'rows' not in report_data or not report_data['rows']:
            return jsonify({'error': 'Нет данных для формирования отчета'}), 404
        
        # Проверяем отмену после получения данных
        if check_if_cancelled():
            return jsonify({'cancelled': True}), 200
            
        try:
            # Считаем только позиции с продажами и вариантами
            total_items = sum(1 for item in report_data['rows'] 
                            if item.get('sellQuantity', 0) > 0 
                            and ('/variant/' in item.get('assortment', {}).get('meta', {}).get('href', '') 
                                or '/product/' in item.get('assortment', {}).get('meta', {}).get('href', '')))
            
            print(f"Всего позиций для обработки: {total_items}")
            
            with processing_lock:
                current_status['total'] = total_items
                current_status['processed'] = 0
            
            excel_file = create_excel_report(report_data, store_id, start_date, end_date, planning_days)
            
            if excel_file is None:
                return jsonify({'cancelled': True}), 200
                
            return jsonify({
                'success': True,
                'file_url': f'/download/{excel_file}'
            })
            
        except Exception as e:
            print(f"Ошибка при обработке данных: {str(e)}")
            import traceback
            print("Полный стек ошибки:")
            print(traceback.format_exc())
            return jsonify({'error': str(e)}), 500
            
    except Exception as e:
        print(f"Общая ошибка в process: {str(e)}")
        import traceback
        print("Полный стек ошибки:")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/cancel', methods=['POST'])
def cancel_processing():
    global processing_cancelled
    with processing_lock:
        processing_cancelled = True
    return jsonify({'status': 'cancelled', 'cancelled': True}), 200

@app.route('/status-stream')
def status_stream():
    def generate():
        last_processed = -1
        while True:
            with processing_lock:
                if current_status['total'] == 0:  # Ждем инициализации
                    yield f"data: ...\n\n"
                else:
                    remaining = current_status['total'] - current_status['processed']
                    if current_status['processed'] != last_processed:
                        last_processed = current_status['processed']
                        # Вычисляем среднее время запроса
                        avg_time = sum(api_request_times) / len(api_request_times) if api_request_times else 0.9
                        status_data = {
                            'remaining': remaining,
                            'processed': current_status['processed'],
                            'total': current_status['total'],
                            'avg_request_time': avg_time
                        }
                        yield f"data: {json.dumps(status_data)}\n\n"
                    if remaining <= 0:
                        break
            sleep(0.1)
    return Response(generate(), mimetype='text/event-stream')

def check_if_cancelled():
    """Проверяет, не была ли отменена обработка отчета"""
    global processing_cancelled
    if processing_cancelled:
        return True
    return False

def get_report_data(start_date, end_date, store_id, product_groups):
    """Получает данные отчета из МойСклад"""
    url = f"{BASE_URL}/report/profit/byvariant"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    # Форматируем даты
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
    
    params = {
        'momentFrom': start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
        'momentTo': end_datetime.replace(hour=23, minute=59, second=59).strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Добавляем фильтры
    filter_parts = []
    if store_id:
        filter_parts.append(f'store={BASE_URL}/entity/store/{store_id}')
    
    if product_groups:
        for group_id in product_groups:
            if group_id:
                filter_parts.append(f'productFolder={BASE_URL}/entity/productfolder/{group_id}')
    
    if filter_parts:
        params['filter'] = ';'.join(filter_parts)
    
    try:
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Ошибка при получении отчета: {response.status_code}")
            return None
    except Exception as e:
        print(f"Ошибка при получении отчета: {str(e)}")
        return None

def get_stores():
    """Получает список складов из МойСклад"""
    url = f"{BASE_URL}/entity/store"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            stores = response.json()['rows']
            return [{'id': store['id'], 'name': store['name']} for store in stores]
        else:
            error_message = f"Ошибка при получении списка складов: {response.status_code}"
            print(error_message)
            return []
    except Exception as e:
        print(f"Ошибка при получении списка складов: {str(e)}")
        return []

def get_product_groups():
    """Получает список групп товаров из МойСклад"""
    url = f"{BASE_URL}/entity/productfolder"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    try:
        all_groups = []
        offset = 0
        limit = 1000

        while True:
            params = {
                'offset': offset,
                'limit': limit,
                'expand': 'productFolder'
            }
            
            response = requests.get(url, headers=headers, params=params)
            if response.status_code == 200:
                data = response.json()
                all_groups.extend(data['rows'])
                if len(data['rows']) < limit:
                    break
                offset += limit
            else:
                error_message = f"Ошибка при получении списка групп товаров: {response.status_code}"
                print(error_message)
                return []

        return build_group_hierarchy(all_groups)
    except Exception as e:
        print(f"Ошибка при получении списка групп товаров: {str(e)}")
        return []

def build_group_hierarchy(groups):
    """Строит иерархию групп товаров"""
    group_dict = {}
    root_groups = []

    # Первый проход - создаем словарь всех групп
    for group in groups:
        group_id = group['id']
        group_dict[group_id] = {
            'id': group_id,
            'name': group['name'],
            'children': [],
            'parent': None,
            'level': 0,
            'has_children': False
        }

    # Второй проход - строим иерархию
    for group in groups:
        group_id = group['id']
        parent_folder = group.get('productFolder', {})
        
        if parent_folder and parent_folder.get('meta'):
            parent_id = parent_folder['meta']['href'].split('/')[-1]
            
            if parent_id in group_dict:
                group_dict[group_id]['parent'] = parent_id
                group_dict[parent_id]['children'].append(group_dict[group_id])
                group_dict[parent_id]['has_children'] = True
        else:
            root_groups.append(group_dict[group_id])

    # Сортируем группы
    root_groups.sort(key=lambda x: x['name'])
    
    def set_levels(groups, level=0):
        for group in groups:
            group['level'] = level
            if group['children']:
                group['has_children'] = True
                set_levels(group['children'], level + 1)
                group['children'].sort(key=lambda x: x['name'])

    set_levels(root_groups)
    return root_groups

def render_group_options(groups, level=0):
    """Рендерит HTML-опции для групп товаров"""
    result = []
    for group in groups:
        indent = '—' * level
        has_children = '1' if group.get('children') and len(group['children']) > 0 else '0'
        
        option_html = (
            f'<option value="{group["id"]}" '
            f'data-level="{level}" '
            f'data-has-children="{has_children}" '
            f'data-parent="{group.get("parent", "")}" '
            f'style="margin-left: {level * 20}px">'
            f'{indent} {group["name"]}'
            f'</option>'
        )
        
        result.append(option_html)
        
        if group.get('children'):
            child_options = render_group_options(group['children'], level + 1)
            result.extend(child_options)
    
    return '\n'.join(result)

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_file(filename, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def truncate_filename(filename, max_length=100):
    """Обрезает имя файла до указанной длины, сохраняя расширение"""
    if len(filename) <= max_length:
        return filename
    
    name, ext = os.path.splitext(filename)
    return name[:max_length-len(ext)] + ext

def get_sales_speed_v2(variant_id, store_id, start_date, end_date, is_variant):
    """Рассчитывает скорость продаж для варианта или товара"""
    print(f"\nНачало расчета скорости продаж v2 для варианта {variant_id}")
    total_start_time = time()
    
    url = f"{BASE_URL}/report/turnover/byoperations"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    # Преобразуем даты в datetime объекты
    date_conversion_start = time()
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
    
    # Создаем дату для поиска предыдущих продаж (100 дней до начала периода)
    extended_start_datetime = start_datetime - timedelta(days=100)
    
    start_date_formatted = extended_start_datetime.strftime('%Y-%m-%d %H:%M:%S')
    end_date_formatted = end_datetime.strftime('%Y-%m-%d %H:%M:%S')
    
    params = {
        'momentFrom': start_date_formatted,
        'momentTo': end_date_formatted,
        'limit': 1000,
        'order': 'moment,desc'
    }
    
    assortment_type = 'variant' if is_variant else 'product'
    filter_params = [
        f"filter=store={BASE_URL}/entity/store/{store_id}",
        f"filter={assortment_type}={BASE_URL}/entity/{assortment_type}/{variant_id}"
    ]
    
    query_string = '&'.join([f"{k}={v}" for k, v in params.items()] + filter_params)
    full_url = f"{url}?{query_string}"
    
    # Замер времени API запроса
    api_request_start = time()
    try:
        response = requests.get(full_url, headers=headers, timeout=30)
    except requests.exceptions.Timeout:
        print(f"Timeout при запросе операций для варианта {variant_id}")
        return 0, '', '', '', '', ''
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при запросе операций для варианта {variant_id}: {str(e)}")
        return 0, '', '', '', '', ''
    
    api_request_time = time() - api_request_start
    print(f"Время выполнения API запроса: {api_request_time:.3f} сек")
    api_request_times.append(api_request_time)
    
    if response.status_code != 200:
        print(f"Ошибка при получении данных: {response.status_code}")
        return 0, '', '', '', '', ''

    data = response.json()
    if not data or 'rows' not in data or not data['rows']:
        return 0, '', '', '', '', ''
    
    # Получаем имя из первой строки ответа
    assortment_name = data['rows'][0]['assortment']['name']
    
    # Получаем все продажи и конвертируем даты один раз
    all_sales = []
    for row in data.get('rows', []):
        if (row['assortment']['meta']['href'].split('/')[-1] == variant_id and
            row['operation']['meta']['type'] == 'retaildemand' and
            row['quantity'] < 0):  # Продажи имеют отрицательное количество
            sale_date = datetime.fromisoformat(row['operation']['moment'].replace('Z', '+00:00'))
            all_sales.append((sale_date, row))
    
    # Сортируем все продажи по дате (по убыванию)
    all_sales.sort(key=lambda x: x[0], reverse=True)
    
    # Находим продажи в указанном периоде
    sales_in_period = [(date, row) for date, row in all_sales 
                      if start_datetime <= date <= end_datetime]
    
    if not sales_in_period:
        return 0, '', '', '', '', assortment_name
    
    # Получение метаданных
    group_uuid = ''
    group_name = ''
    product_uuid = ''
    product_href = ''
    
    assortment = sales_in_period[0][1].get('assortment', {})
    product_folder = assortment.get('productFolder', {})
    group_href = product_folder.get('meta', {}).get('href', '')
    group_uuid = group_href.split('/')[-1] if group_href else ''
    group_name = product_folder.get('name', '')
    
    product_meta = assortment.get('meta', {})
    product_href = product_meta.get('uuidHref', '')
    if product_href:
        product_uuid = product_meta.get('href', '').split('/')[-1]
    
    # Находим date2 - последняя продажа в периоде
    date2 = sales_in_period[0][0]
    
    # Находим date1 - последняя продажа до начала периода
    date1 = start_datetime
    
    # Ищем последнюю продажу до начала периода
    found_sale_before_period = False
    for date, _ in all_sales:
        if date < start_datetime:
            date1 = date
            found_sale_before_period = True
            break
    
    # Вычисляем разницу в днях
    days = (date2 - date1).total_seconds() / (24 * 60 * 60)
    
    # Считаем общее количество проданных единиц за период
    total_sold = sum(abs(row['quantity']) for _, row in sales_in_period)
    
    if days == 0:
        return total_sold, group_uuid, group_name, product_uuid, product_href, assortment_name
    
    # Вычисляем среднюю скорость продаж
    sales_speed = total_sold / days
    
    return sales_speed, group_uuid, group_name, product_uuid, product_href, assortment_name

def create_excel_report(report_data, store_id, start_date, end_date, planning_days):
    """Создает Excel отчет на основе полученных данных"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Анализ1"
        
        # Заголовки для первого листа
        headers = [
            "Наименование", "Артикул", "Код", "Группа", "Продажи шт.", 
            "Прибыль", "Себестоимость", "Выручка", "Рентабельность",
            "Средняя скорость продаж", "Прогноз продаж", "Прогноз прибыли"
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем границы для заголовков
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Устанавливаем ширину столбцов
        column_widths = {
            1: 40,  # Наименование
            2: 15,  # Артикул
            3: 15,  # Код
            4: 30,  # Группа
            5: 15,  # Продажи шт.
            6: 15,  # Прибыль
            7: 15,  # Себестоимость
            8: 15,  # Выручка
            9: 15,  # Рентабельность
            10: 15, # Средняя скорость продаж
            11: 15, # Прогноз продаж
            12: 15  # Прогноз прибыли
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Создаем второй лист для анализа по группам
        ws2 = wb.create_sheet("Анализ2")
        
        # Заголовки для второго листа
        group_headers = [
            "Группа", "Количество SKU", "Продажи шт.", 
            "Прибыль", "Себестоимость", "Выручка", "Рентабельность",
            "Прогноз продаж", "Прогноз прибыли"
        ]
        
        # Записываем заголовки на втором листе
        for col, header in enumerate(group_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Устанавливаем ширину столбцов на втором листе
        group_column_widths = {
            1: 40,  # Группа
            2: 15,  # Количество SKU
            3: 15,  # Продажи шт.
            4: 15,  # Прибыль
            5: 15,  # Себестоимость
            6: 15,  # Выручка
            7: 15,  # Рентабельность
            8: 15,  # Прогноз продаж
            9: 15   # Прогноз прибыли
        }
        
        for col, width in group_column_widths.items():
            ws2.column_dimensions[get_column_letter(col)].width = width
        
        # Словарь для хранения данных по группам
        group_data = {}
        
        # Обработка данных и заполнение первого листа
        row = 2
        for item in report_data['rows']:
            # Проверяем наличие продаж
            if item.get('sellQuantity', 0) <= 0:
                continue
                
            # Получаем мета-информацию о товаре/варианте
            assortment = item.get('assortment', {})
            meta = assortment.get('meta', {})
            href = meta.get('href', '')
            
            # Пропускаем, если это не товар и не вариант
            if '/variant/' not in href and '/product/' not in href:
                continue
            
            # Проверяем отмену
            if check_if_cancelled():
                wb.close()
                return None
            
            # Получаем ID и определяем тип (товар или вариант)
            entity_id = href.split('/')[-1]
            is_variant = '/variant/' in href
            
            # Рассчитываем скорость продаж
            sales_speed, group_uuid, group_name, product_uuid, product_href, assortment_name = get_sales_speed_v2(
                entity_id, store_id, start_date, end_date, is_variant
            )
            
            # Обновляем счетчик обработанных позиций
            with processing_lock:
                current_status['processed'] += 1
            
            # Основные показатели
            sell_quantity = abs(item.get('sellQuantity', 0))
            profit = item.get('profit', 0)
            cost_price = item.get('costPrice', 0)
            revenue = item.get('revenue', 0)
            
            # Рассчитываем рентабельность
            profitability = (profit / revenue * 100) if revenue > 0 else 0
            
            # Прогнозы
            sales_forecast = math.ceil(sales_speed * planning_days)
            profit_forecast = (profit / sell_quantity * sales_forecast) if sell_quantity > 0 else 0
            
            # Записываем данные в строку
            ws.cell(row=row, column=1, value=assortment_name)
            ws.cell(row=row, column=2, value=assortment.get('article', ''))
            ws.cell(row=row, column=3, value=assortment.get('code', ''))
            ws.cell(row=row, column=4, value=group_name)
            ws.cell(row=row, column=5, value=sell_quantity)
            ws.cell(row=row, column=6, value=profit)
            ws.cell(row=row, column=7, value=cost_price)
            ws.cell(row=row, column=8, value=revenue)
            ws.cell(row=row, column=9, value=profitability)
            ws.cell(row=row, column=10, value=sales_speed)
            ws.cell(row=row, column=11, value=sales_forecast)
            ws.cell(row=row, column=12, value=profit_forecast)
            
            # Форматирование ячеек
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Выравнивание
                if col == 1:  # Наименование
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Форматирование чисел
                if col in [6, 7, 8, 12]:  # Денежные значения
                    cell.number_format = '#,##0.00'
                elif col in [9]:  # Проценты
                    cell.number_format = '0.00%'
                elif col in [10]:  # Скорость продаж
                    cell.number_format = '0.00'
            
            # Добавляем данные в группу
            if group_uuid not in group_data:
                group_data[group_uuid] = {
                    'name': group_name,
                    'sku_count': 0,
                    'sell_quantity': 0,
                    'profit': 0,
                    'cost_price': 0,
                    'revenue': 0,
                    'sales_forecast': 0,
                    'profit_forecast': 0
                }
            
            group_data[group_uuid]['sku_count'] += 1
            group_data[group_uuid]['sell_quantity'] += sell_quantity
            group_data[group_uuid]['profit'] += profit
            group_data[group_uuid]['cost_price'] += cost_price
            group_data[group_uuid]['revenue'] += revenue
            group_data[group_uuid]['sales_forecast'] += sales_forecast
            group_data[group_uuid]['profit_forecast'] += profit_forecast
            
            row += 1
        
        # Заполняем второй лист данными по группам
        row = 2
        for group_uuid, data in group_data.items():
            # Рассчитываем рентабельность для группы
            group_profitability = (data['profit'] / data['revenue'] * 100) if data['revenue'] > 0 else 0
            
            # Записываем данные группы
            ws2.cell(row=row, column=1, value=data['name'])
            ws2.cell(row=row, column=2, value=data['sku_count'])
            ws2.cell(row=row, column=3, value=data['sell_quantity'])
            ws2.cell(row=row, column=4, value=data['profit'])
            ws2.cell(row=row, column=5, value=data['cost_price'])
            ws2.cell(row=row, column=6, value=data['revenue'])
            ws2.cell(row=row, column=7, value=group_profitability)
            ws2.cell(row=row, column=8, value=data['sales_forecast'])
            ws2.cell(row=row, column=9, value=data['profit_forecast'])
            
            # Форматирование ячеек
            for col in range(1, 10):
                cell = ws2.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Выравнивание
                if col == 1:  # Название группы
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Форматирование чисел
                if col in [4, 5, 6, 9]:  # Денежные значения
                    cell.number_format = '#,##0.00'
                elif col in [7]:  # Проценты
                    cell.number_format = '0.00%'
            
            row += 1
        
        # Добавляем строку с итогами на втором листе
        row = ws2.max_row + 1
        ws2.cell(row=row, column=1, value="ИТОГО:")
        ws2.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для суммирования
        for col in range(2, 10):
            column_letter = get_column_letter(col)
            cell = ws2.cell(row=row, column=col)
            cell.font = Font(bold=True)
            
            if col != 7:  # Все колонки кроме рентабельности
                cell.value = f"=SUM({column_letter}2:{column_letter}{row-1})"
            else:  # Рентабельность
                cell.value = f"=D{row}/F{row}"
                cell.number_format = '0.00%'
            
            # Копируем форматирование из предыдущей ячейки
            source_cell = ws2.cell(row=row-1, column=col)
            if source_cell.number_format:
                cell.number_format = source_cell.number_format
            
            # Добавляем границы
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Выравнивание
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Копируем цвет фона
            source_cell = ws2.cell(row=row-1, column=col)
            if source_cell.fill and hasattr(source_cell.fill, 'start_color') and source_cell.fill.start_color:
                fill_color = source_cell.fill.start_color.rgb or 'FFFFFF'
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )
        
        # Создаем лист с информацией
        info_ws = wb.create_sheet("Инфо")
        
        # Получаем название магазина
        stores = get_stores()
        store_name = next((store['name'] for store in stores if store['id'] == store_id), store_id)
        
        # Форматируем даты
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        start_date_formatted = start_date_obj.strftime('%d.%m.%Y')
        end_date_formatted = end_date_obj.strftime('%d.%m.%Y')
        
        # Дата за 100 дней до начала периода
        extended_start_date = start_date_obj - timedelta(days=100)
        extended_start_formatted = extended_start_date.strftime('%d.%m.%Y')
        
        # Информация для листа
        info_data = [
            ["Параметры анализа", ""],
            ["", ""],
            ["Период анализа:", f"с {start_date_formatted} по {end_date_formatted}"],
            ["Дата начала поиска последней продажи перед Дата начала:", extended_start_formatted],
            ["Количество дней анализа:", f"{(end_date_obj - start_date_obj).days + 1} дней"],
            ["", ""],
            ["Магазин:", store_name],
            ["", ""],
            ["Период планирования:", f"{planning_days} дней"],
            ["", ""],
            ["Выбранные группы товаров:", ""],
        ]
        
        # Добавляем информацию о выбранных группах
        product_groups = []
        if 'final_product_groups' in request.form:
            raw_groups = request.form.get('final_product_groups', '')
            if raw_groups:
                product_groups = [group.strip() for group in raw_groups.split(',') if group.strip()]
        
        # Получаем пути групп
        raw_paths = request.form.get('final_product_paths', '')
        if raw_paths:
            selected_groups = []
            second_level_names = []
            for path in raw_paths.split('||'):
                if path.strip():
                    clean_path_parts = [part.strip() for part in path.strip().split('/')]
                    selected_groups.append('/'.join(clean_path_parts))
                    
                    if len(clean_path_parts) >= 2:
                        if len(clean_path_parts) == 2 or (len(clean_path_parts) > 2 and clean_path_parts[2] == "Выберите подгруппу"):
                            second_level_names.append(clean_path_parts[1])
                        elif len(clean_path_parts) > 2:
                            second_level_names.append(clean_path_parts[2])
        
        # Формируем имя файла
        group_name_for_file = "Все группы"
        if second_level_names:
            group_name_for_file = ','.join(second_level_names)
        
        # Добавляем выбранные группы в информационный лист
        if selected_groups:
            for group_path in selected_groups:
                info_data.append(["", group_path])
        else:
            info_data.append(["", "Все группы"])
        
        # Записываем данные в лист
        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = info_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (col_idx == 1 and value):
                    cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical='center')
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Устанавливаем ширину столбцов
        info_ws.column_dimensions['A'].width = 30
        info_ws.column_dimensions['B'].width = 60
        
        # Добавляем время создания отчета
        current_time = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        time_row = len(info_data) + 2
        info_ws.cell(row=time_row, column=1, value="Отчет сформирован:").font = Font(bold=True, size=10)
        info_ws.cell(row=time_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        info_ws.cell(row=time_row, column=2, value=current_time).font = Font(size=10)
        info_ws.cell(row=time_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
        
        # Делаем активным первый лист
        wb.active = wb["Анализ1"]
        
        # Формируем имя файла
        filename = f"{start_date_formatted}-{end_date_formatted} - {store_name} - {group_name_for_file}.xlsx"
        filename = truncate_filename(filename)
        
        wb.save(filename)
        wb.close()
        return filename
        
    except Exception as e:
        try:
            wb.close()
        except:
            pass
        raise e

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=port)
